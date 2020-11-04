# open_workbook.py

from openpyxl import load_workbook


def open_workbook(path):
    workbook = load_workbook(filename=path)
    print(f'Worksheet names: {workbook.sheetnames}')
    sheet = workbook.active
    print(sheet)
    print(f'The title of the Worksheet is: {sheet.title}')


def get_cell_info(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    print(sheet)
    print(f'The title of the Worksheet is: {sheet.title}')
    print(f'The value of {sheet["K1"].value=}')
    print(f'The value of {sheet["K2"].value=}')
    cell = sheet['K3']
    print(f'{cell.value=}')


def get_info_by_coord(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    cell = sheet['K2']
    print(f'Row {cell.row}, Col {cell.column} = {cell.value}')
    print(f'{cell.value=} is at {cell.coordinate=}')


if __name__ == '__main__':
    open_workbook('/Users/WRA1523/OneDrive - Emerson/Varsity/!Archive/SP 743 - Why Orders with a Stop Code of 900/Derby Orders - 2020.10.19(0).xlsx')
    print("Next:\t")
    get_cell_info('/Users/WRA1523/OneDrive - Emerson/Varsity/!Archive/SP 743 - Why Orders with a Stop Code of 900/Derby Orders - 2020.10.19(0).xlsx')
    print("Next 2:\t")
    get_info_by_coord('/Users/WRA1523/OneDrive - Emerson/Varsity/!Archive/SP 743 - Why Orders with a Stop Code of 900/Derby Orders - 2020.10.19(0).xlsx')
