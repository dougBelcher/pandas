# remove_sheets.py

import openpyxl


def remove_worksheets(path):
    workbook = openpyxl.Workbook()
    sheet1 = workbook.create_sheet()
    # Insert a worksheet
    workbook.create_sheet(index=1,
                          title='Second sheet')
    print(workbook.sheetnames)
    workbook.remove(sheet1)
    print(workbook.sheetnames)
    workbook.save(path)


if __name__ == '__main__':
    remove_worksheets('/Users/WRA1523/OneDrive - Emerson/Varsity/!Archive/remove_sheets.xlsx')