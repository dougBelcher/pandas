# iterating_over_cells.py

from openpyxl import load_workbook


def iterating_range(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    for cell in sheet['A']:
        print(cell)


def iterating_over_values(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    for value in sheet.iter_rows(
            min_row=1, max_row=3,
            min_col=1, max_col=3,
            values_only=True,
        ):
        print(value)


if __name__ == '__main__':
    # iterating_range('/Users/WRA1523/OneDrive - Emerson/Varsity/!Archive/SP 743 - Why Orders with a Stop Code of 900/Derby Orders - 2020.10.19(0).xlsx')
    # print("Next:\t")
    iterating_over_values('/Users/WRA1523/OneDrive - Emerson/Varsity/!Archive/SP 743 - Why Orders with a Stop Code of 900/Derby Orders - 2020.10.19(0).xlsx')